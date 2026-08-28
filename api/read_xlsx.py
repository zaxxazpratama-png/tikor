#!/usr/bin/env python3
import sys
import json
import openpyxl

def read_xlsx(file_path, sheet_number=2):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        
        # Sheet number is 1-indexed
        sheet_idx = int(sheet_number) - 1
        if sheet_idx < 0 or sheet_idx >= len(sheets):
            return {"error": f"Sheet {sheet_number} tidak ditemukan. Available sheets: {sheets}"}
        
        ws = wb.worksheets[sheet_idx]
        rows = []
        
        for row in ws.iter_rows(values_only=True):
            processed_row = []
            for cell in row:
                if cell is None:
                    processed_row.append(None)
                else:
                    val = str(cell) if cell is not None else None
                    processed_row.append(val)
            rows.append(processed_row)
        
        wb.close()
        return {"rows": rows, "total": len(rows)}
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: read_xlsx.py <file_path> [sheet_number]"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    sheet_number = int(sys.argv[2]) if len(sys.argv) > 2 else 2
    
    result = read_xlsx(file_path, sheet_number)
    print(json.dumps(result, ensure_ascii=False))
